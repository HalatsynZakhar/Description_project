"""Консольна обробка Excel-карток товарів за правилами MONO."""

from __future__ import annotations

import json
import os
import random
import re
import tempfile
import time
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Protocol

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input"
PROMPT_PATH = BASE_DIR / "prompts" / "mono_product_prompt.txt"
KEYS_PATH = BASE_DIR / "keys.json"
KEY_STATE_PATH = BASE_DIR / "keys_state.json"
ERROR_LOG_PATH = BASE_DIR / "logs" / "gemini_errors.jsonl"
MODEL_NAME = "gemini-3.5-flash-lite"
TITLE_HEADER = "Назва MONO"
DESCRIPTION_HEADER = "Опис MONO"
HEADER_ROW = 1
INVALID_KEY_COOLDOWN = timedelta(hours=24)
MAX_TRANSIENT_RETRIES = 2
# Google SDK uses milliseconds. A request that takes longer is treated as a
# temporary API failure, so the next key can be tried instead of waiting forever.
HTTP_TIMEOUT_MS = 120_000

ALLOWED_HTML_TAGS = {"h5", "br", "p", "ul", "li"}
HTML_TAG_PATTERN = re.compile(r"</?\s*([a-zA-Z0-9]+)(?:\s+[^<>]*)?\s*/?>")
DECORATIVE_TAG_PATTERN = re.compile(r"</?\s*(?:strong|b|em|i)\s*>", re.IGNORECASE)
URL_PATTERN = re.compile(r"(?:https?://|www\.)", re.IGNORECASE)
BANNED_TITLE_WORDS = {
    "акція",
    "знижка",
    "розпродаж",
    "уцінка",
    "copy",
    "original",
}
EXCLUDED_SOURCE_SECTIONS = {"комплектація", "додаткова інформація"}
BLOCKED_SOURCE_LINE_PATTERN = re.compile(r"tell\s+my\s+fortune", re.IGNORECASE)


class ProcessorError(Exception):
    """Помилка, яку можна показати користувачу без технічного стеку."""


class NoAvailableKeysError(ProcessorError):
    """Усі ключі тимчасово недоступні."""


class GenerationError(ProcessorError):
    """Модель не повернула придатного результату для одного рядка."""


class RetryableGenerationError(GenerationError):
    """Відповідь моделі можна безпечно один раз повторити."""


@dataclass(frozen=True)
class ApiKey:
    name: str
    value: str


@dataclass
class ProcessingSummary:
    processed: int = 0
    skipped_completed: int = 0
    skipped_missing_source: int = 0
    failed_rows: list[int] | None = None

    def __post_init__(self) -> None:
        if self.failed_rows is None:
            self.failed_rows = []


class ProductGenerator(Protocol):
    def generate(
        self,
        source_title: str,
        source_description: str,
        *,
        generate_title: bool = True,
        generate_description: bool = True,
        existing_title: str = "",
        existing_description: str = "",
        article_code: str = "",
    ) -> dict[str, str]: ...


def utc_now() -> datetime:
    return datetime.now(UTC)


def is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def cell_as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def read_json(path: Path, default: dict[str, Any]) -> dict[str, Any]:
    if not path.exists():
        return default
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise ProcessorError(f"Не вдалося прочитати {path.name}: {error}") from error
    if not isinstance(data, dict):
        raise ProcessorError(f"Файл {path.name} має містити JSON-об’єкт.")
    return data


def write_json_atomically(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w", encoding="utf-8", suffix=".tmp", dir=path.parent, delete=False
        ) as temporary_file:
            temporary_path = Path(temporary_file.name)
            json.dump(data, temporary_file, ensure_ascii=False, indent=2)
            temporary_file.flush()
            os.fsync(temporary_file.fileno())
        os.replace(temporary_path, path)
    finally:
        if temporary_path and temporary_path.exists():
            temporary_path.unlink(missing_ok=True)


def log_gemini_error(event: str, **details: Any) -> None:
    """Дописує діагностику без API-ключів і без текстів товарів."""
    record = {
        "timestamp": utc_now().isoformat(),
        "event": event,
        **details,
    }
    ERROR_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    with ERROR_LOG_PATH.open("a", encoding="utf-8", newline="\n") as file:
        json.dump(record, file, ensure_ascii=False)
        file.write("\n")
        file.flush()
        os.fsync(file.fileno())


class KeyPool:
    """Ключі Gemini зі збереженим станом тимчасової недоступності."""

    def __init__(self, keys_path: Path = KEYS_PATH, state_path: Path = KEY_STATE_PATH) -> None:
        raw_keys = read_json(keys_path, {"keys": []}).get("keys", [])
        if not isinstance(raw_keys, list):
            raise ProcessorError('Поле "keys" у keys.json має бути списком.')

        self.keys: list[ApiKey] = []
        names: set[str] = set()
        for position, item in enumerate(raw_keys, start=1):
            if not isinstance(item, dict) or item.get("enabled", True) is False:
                continue
            name = str(item.get("name", f"key-{position}")).strip()
            value = str(item.get("api_key", "")).strip()
            if not name or not value or value.startswith("ВСТАВТЕ_"):
                continue
            if name in names:
                raise ProcessorError(f'У keys.json повторюється ім’я ключа "{name}".')
            names.add(name)
            self.keys.append(ApiKey(name=name, value=value))

        if not self.keys:
            raise ProcessorError(
                "Не знайдено активних ключів. Заповніть keys.json за прикладом keys.example.json."
            )

        self.state_path = state_path
        self.state = read_json(state_path, {"version": 1, "keys": {}})
        if not isinstance(self.state.get("keys"), dict):
            self.state["keys"] = {}

    def _record(self, key: ApiKey) -> dict[str, Any]:
        records = self.state["keys"]
        record = records.get(key.name)
        if not isinstance(record, dict):
            record = {}
            records[key.name] = record
        return record

    @staticmethod
    def _parse_time(value: Any) -> datetime | None:
        if not isinstance(value, str):
            return None
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return None

    def available(self) -> Iterable[ApiKey]:
        now = utc_now()
        for key in self.keys:
            blocked_until = self._parse_time(self._record(key).get("blocked_until"))
            if blocked_until is None or blocked_until <= now:
                yield key

    def mark_success(self, key: ApiKey) -> None:
        record = self._record(key)
        record.update(
            {
                "blocked_until": None,
                "reason": None,
                "last_error": None,
                "failure_count": 0,
                "last_success_at": utc_now().isoformat(),
            }
        )
        write_json_atomically(self.state_path, self.state)

    def mark_failure(
        self, key: ApiKey, category: str, message: str, retry_after_seconds: float | None = None
    ) -> datetime:
        record = self._record(key)
        failures = int(record.get("failure_count", 0)) + 1
        if category == "invalid_key":
            cooldown = INVALID_KEY_COOLDOWN
        elif retry_after_seconds is not None:
            cooldown = timedelta(seconds=max(1, min(retry_after_seconds, 12 * 60 * 60)))
        else:
            seconds = min(15 * 60, 30 * (2 ** min(failures - 1, 5)))
            cooldown = timedelta(seconds=seconds + random.uniform(0, 3))
        blocked_until = utc_now() + cooldown
        record.update(
            {
                "blocked_until": blocked_until.isoformat(),
                "reason": category,
                "last_error": message[:500],
                "failure_count": failures,
                "last_failure_at": utc_now().isoformat(),
            }
        )
        write_json_atomically(self.state_path, self.state)
        return blocked_until

    def next_available_time(self) -> datetime | None:
        blocked_times = [
            self._parse_time(self._record(key).get("blocked_until")) for key in self.keys
        ]
        future_times = [moment for moment in blocked_times if moment and moment > utc_now()]
        return min(future_times) if future_times else None


def load_prompt() -> str:
    try:
        return PROMPT_PATH.read_text(encoding="utf-8")
    except OSError as error:
        raise ProcessorError(f"Не вдалося прочитати промпт {PROMPT_PATH}: {error}") from error


def normalized_section_heading(line: str) -> str:
    """Повертає заголовок секції без HTML, двокрапки та зайвих пробілів."""
    plain = strip_html(line).strip().lower()
    return plain.rstrip(":").strip()


def source_description_for_model(source_description: str) -> str:
    """Викидає секції, які вимоги MONO прямо забороняють переносити в опис."""
    kept_lines: list[str] = []
    skip_section = False
    for line in source_description.splitlines():
        if BLOCKED_SOURCE_LINE_PATTERN.search(line):
            continue
        heading = normalized_section_heading(line)
        if heading in EXCLUDED_SOURCE_SECTIONS:
            skip_section = True
            continue
        # Зустріли новий короткий заголовок секції — знову читаємо потрібні дані.
        is_heading = bool(heading) and line.strip().endswith(":") and not line.lstrip().startswith("*")
        if skip_section and is_heading:
            skip_section = False
        if not skip_section:
            kept_lines.append(line)
    cleaned = "\n".join(kept_lines).strip()
    return cleaned or source_description


def response_diagnostics(response: Any) -> str:
    """Коротка причина порожньої відповіді без збереження даних товару."""
    candidates = getattr(response, "candidates", None) or []
    reasons = [
        str(getattr(candidate, "finish_reason", ""))
        for candidate in candidates
        if getattr(candidate, "finish_reason", None)
    ]
    feedback = getattr(response, "prompt_feedback", None)
    block_reason = getattr(feedback, "block_reason", None) if feedback else None
    parts = []
    if reasons:
        parts.append("finish_reason=" + ",".join(reasons))
    if block_reason:
        parts.append(f"block_reason={block_reason}")
    return "; ".join(parts) or "причину не надано API"


class GeminiGenerator:
    """Один виклик Gemini повертає і назву, і опис товару."""

    def __init__(self, pool: KeyPool, prompt_template: str | None = None) -> None:
        self.pool = pool
        self.prompt_template = prompt_template or load_prompt()

    def _request(self, api_key: str, prompt: str, fields: list[str]) -> dict[str, str]:
        try:
            from google import genai
            from google.genai import types
        except ImportError as error:
            raise ProcessorError(
                "Не встановлено залежності. Виконайте: .\\.venv\\Scripts\\python.exe -m pip install -r requirements.txt"
            ) from error

        client = genai.Client(
            api_key=api_key,
            http_options=types.HttpOptions(
                timeout=HTTP_TIMEOUT_MS,
                # The application owns retries and key rotation. SDK default is 5,
                # which can make one invisible request wait for several minutes.
                retryOptions=types.HttpRetryOptions(attempts=1),
            ),
        )
        response = client.models.generate_content(
            model=MODEL_NAME,
            contents=prompt,
            config=types.GenerateContentConfig(
                temperature=0.2,
                response_mime_type="application/json",
                response_schema={
                    "type": "object",
                    "properties": {field: {"type": "string"} for field in fields},
                    "required": fields,
                },
            ),
        )
        text = getattr(response, "text", None)
        if not text:
            raise RetryableGenerationError(
                "Gemini не повернув текст відповіді (" + response_diagnostics(response) + ")."
            )
        try:
            data = json.loads(text)
        except json.JSONDecodeError as error:
            raise RetryableGenerationError("Gemini повернув не JSON-відповідь.") from error
        if not isinstance(data, dict):
            raise GenerationError("Відповідь Gemini має бути JSON-об’єктом.")
        return {field: cell_as_text(data.get(field)) for field in fields}

    def generate(
        self,
        source_title: str,
        source_description: str,
        *,
        generate_title: bool = True,
        generate_description: bool = True,
        existing_title: str = "",
        existing_description: str = "",
        article_code: str = "",
    ) -> dict[str, str]:
        fields = []
        if generate_title:
            fields.append("title")
        if generate_description:
            fields.append("description")
        if not fields:
            raise ProcessorError("Немає полів для генерації.")

        if fields == ["title"]:
            generation_task = "Створи ЛИШЕ українську назву. Поле description не повертай."
        elif fields == ["description"]:
            generation_task = "Створи ЛИШЕ український опис. Поле title не повертай."
        else:
            generation_task = "Створи одночасно українську назву й український опис."
        response_contract = json.dumps({field: "рядок" for field in fields}, ensure_ascii=False)
        source_for_model = source_description_for_model(source_description)
        prompt = (
            self.prompt_template.replace("{response_contract}", response_contract)
            .replace("{generation_task}", generation_task)
            .replace("{source_title}", source_title)
            .replace("{source_description}", source_for_model)
        )
        if existing_title:
            prompt += (
                "\n\nУже затверджена назва MONO (лише контекст, не повертай її):\n"
                + existing_title
            )
        if existing_description:
            prompt += (
                "\n\nУже затверджений опис MONO (лише контекст, не повертай його):\n"
                + existing_description
            )
        if article_code and generate_title:
            prompt += (
                "\n\nАртикул буде додано програмою в кінці назви: ("
                + article_code
                + "). Не додавай цей артикул до поля title самостійно."
            )
        last_error: Exception | None = None
        for key in self.pool.available():
            for attempt in range(MAX_TRANSIENT_RETRIES):
                try:
                    print(
                        f"Gemini: ключ «{key.name}», спроба {attempt + 1}/{MAX_TRANSIENT_RETRIES}…",
                        flush=True,
                    )
                    result = self._request(key.value, prompt, fields)
                    validate_result(
                        result,
                        source_description,
                        validate_title_field=generate_title,
                        validate_description_field=generate_description,
                        article_code=article_code,
                    )
                    self.pool.mark_success(key)
                    return result
                except RetryableGenerationError as error:
                    last_error = error
                    log_gemini_error(
                        "retryable_model_response",
                        key_name=key.name,
                        attempt=attempt + 1,
                        error_type=type(error).__name__,
                        error_message=str(error),
                    )
                    if attempt + 1 < MAX_TRANSIENT_RETRIES:
                        print("Gemini: неповна відповідь, повторюю запит…", flush=True)
                        time.sleep((2**attempt) + random.uniform(0, 0.5))
                        continue
                    raise
                except GenerationError:
                    # Відповідь моделі не є ознакою несправного ключа.
                    raise
                except Exception as error:  # API SDK має різні класи винятків у різних версіях.
                    last_error = error
                    category, retry_after = classify_api_error(error)
                    if category == "request_error":
                        log_gemini_error(
                            "api_request_error",
                            key_name=key.name,
                            attempt=attempt + 1,
                            error_type=type(error).__name__,
                            error_message=str(error),
                        )
                        raise ProcessorError(f"Помилка запиту Gemini: {error}") from error
                    if category == "transient" and attempt + 1 < MAX_TRANSIENT_RETRIES:
                        print("Gemini: тимчасова помилка, повторюю запит…", flush=True)
                        time.sleep((2**attempt) + random.uniform(0, 0.5))
                        continue
                    self.pool.mark_failure(key, category, str(error), retry_after)
                    log_gemini_error(
                        "api_unavailable",
                        key_name=key.name,
                        attempt=attempt + 1,
                        category=category,
                        retry_after_seconds=retry_after,
                        error_type=type(error).__name__,
                        error_message=str(error),
                    )
                    print(
                        f"Gemini: ключ «{key.name}» тимчасово відкладено ({category}).",
                        flush=True,
                    )
                    break

        next_time = self.pool.next_available_time()
        when = next_time.astimezone().strftime("%d.%m.%Y %H:%M:%S") if next_time else "невідомо"
        raise NoAvailableKeysError(
            f"Немає доступних ключів Gemini. Наступна перевірка ключа: {when}. "
            f"Остання помилка: {last_error}"
        )


def classify_api_error(error: Exception) -> tuple[str, float | None]:
    """Повертає категорію помилки та, якщо можливо, Retry-After у секундах."""
    status = getattr(error, "status_code", None) or getattr(error, "code", None)
    text = str(error)
    if not isinstance(status, int):
        found = re.search(r"\b(401|403|408|429|500|501|502|503|504)\b", text)
        status = int(found.group(1)) if found else None

    retry_match = re.search(r"retry (?:in|after)\s+([0-9.]+)\s*s", text, re.IGNORECASE)
    retry_after = float(retry_match.group(1)) if retry_match else None
    if status in {401, 403}:
        return "invalid_key", retry_after
    if status == 429:
        return "rate_limited", retry_after
    if status in {408, 500, 502, 503, 504} or status is None:
        return "transient", retry_after
    return "request_error", retry_after


def strip_html(value: str) -> str:
    return re.sub(r"<[^>]*>", "", value)


def remove_decorative_tags(description: str) -> str:
    """Теги оформлення не дозволені MONO й не змінюють зміст тексту."""
    return DECORATIVE_TAG_PATTERN.sub("", description)


def validate_title(title: str) -> None:
    if not title:
        raise GenerationError("Gemini повернув порожню назву.")
    if len(title) > 100:
        raise GenerationError("Назва довша за 100 символів.")
    if URL_PATTERN.search(title):
        raise GenerationError("У назві знайдено посилання.")
    lower_words = set(re.findall(r"[\w’'-]+", title.lower(), flags=re.UNICODE))
    banned = lower_words & BANNED_TITLE_WORDS
    if banned:
        raise GenerationError(f"У назві є заборонені слова: {', '.join(sorted(banned))}.")
    if re.search(r"[§≠≥]", title):
        raise GenerationError("У назві є заборонений спеціальний символ.")
    letters = "".join(char for char in title if char.isalpha())
    if len(letters) > 8 and letters.isupper():
        raise GenerationError("Назва повністю написана CAPS LOCK.")


def shorten_title(title: str, limit: int = 100) -> str:
    """Скорочує назву лише за пробілом, зберігаючи основні початкові дані."""
    if len(title) <= limit:
        return title
    words = title.split()
    while words and len(" ".join(words)) > limit:
        words.pop()
    shortened = " ".join(words).rstrip(" ,.;:-")
    if not shortened:
        # Крайній випадок: один надто довгий технічний код не можна розрізати мовчки.
        raise GenerationError("Назву неможливо скоротити до 100 символів без розриву слова.")
    return shortened


def append_article_to_title(title: str, article_code: str) -> str:
    """Додає артикул і за потреби скорочує назву з урахуванням цього суфікса."""
    if not article_code:
        return shorten_title(title)
    suffix = f" ({article_code})"
    if len(suffix) >= 100:
        raise GenerationError("Артикул надто довгий: для назви не залишається місця.")
    # Якщо модель попри інструкцію вже додала той самий артикул, дубля не буде.
    if title.endswith(suffix):
        title = title[: -len(suffix)].rstrip()
    shortened = shorten_title(title, limit=100 - len(suffix))
    return shortened + suffix


def validate_description(description: str, source_description: str) -> None:
    if not description:
        raise GenerationError("Gemini повернув порожній опис.")
    if len(description) > 30_000:
        raise GenerationError("Опис перевищує допустимий розмір комірки Excel.")
    if URL_PATTERN.search(description):
        raise GenerationError("В описі знайдено URL або зовнішнє посилання.")
    tags = {match.group(1).lower() for match in HTML_TAG_PATTERN.finditer(description)}
    disallowed = tags - ALLOWED_HTML_TAGS
    if disallowed:
        raise GenerationError(f"В описі є заборонені HTML-теги: {', '.join(sorted(disallowed))}.")
    # Довжину контролює промпт. Відносна межа відхиляла коректні детальні описи.
    # source_description збережено у сигнатурі для сумісності викликів.
    _ = source_description


def validate_result(
    result: dict[str, str],
    source_description: str,
    *,
    validate_title_field: bool = True,
    validate_description_field: bool = True,
    article_code: str = "",
) -> None:
    if validate_title_field:
        original_title = result.get("title", "")
        result["title"] = append_article_to_title(original_title, article_code)
        if result["title"] != original_title:
            print("Gemini: назву скорочено до 100 символів.", flush=True)
        validate_title(result.get("title", ""))
    if validate_description_field:
        result["description"] = remove_decorative_tags(result.get("description", ""))
        validate_description(result["description"], source_description)


def save_workbook_atomically(workbook: Workbook, path: Path) -> None:
    """Зберігає книгу через файл у тій самій теці, не пошкоджуючи оригінал при збої."""
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", dir=path.parent, delete=False) as file:
            temporary_path = Path(file.name)
        workbook.save(temporary_path)
        # На Windows os.fsync потребує дескриптор із правом запису.
        with temporary_path.open("rb+") as file:
            os.fsync(file.fileno())
        os.replace(temporary_path, path)
    except PermissionError as error:
        raise ProcessorError(
            "Не вдається зберегти Excel-файл. Закрийте його в Excel і запустіть програму повторно."
        ) from error
    finally:
        if temporary_path and temporary_path.exists():
            temporary_path.unlink(missing_ok=True)


def headers_in_sheet(sheet: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        value = cell_as_text(sheet.cell(HEADER_ROW, column).value)
        if value and value not in headers:
            headers[value] = column
    return headers


def get_or_create_output_column(sheet: Worksheet, header: str) -> int:
    headers = headers_in_sheet(sheet)
    if header in headers:
        return headers[header]
    column = sheet.max_column + 1
    sheet.cell(HEADER_ROW, column).value = header
    return column


def run_processing(
    workbook_path: Path,
    sheet_name: str,
    source_title_column: int,
    source_description_column: int,
    generator: ProductGenerator,
    article_column: int | None = None,
) -> ProcessingSummary:
    workbook = load_workbook(workbook_path)
    sheet = workbook[sheet_name]
    title_output_column = get_or_create_output_column(sheet, TITLE_HEADER)
    description_output_column = get_or_create_output_column(sheet, DESCRIPTION_HEADER)
    # Заголовки також мають пережити раптове завершення до першого запиту.
    save_workbook_atomically(workbook, workbook_path)

    summary = ProcessingSummary()
    for row in range(HEADER_ROW + 1, sheet.max_row + 1):
        source_title = cell_as_text(sheet.cell(row, source_title_column).value)
        source_description = cell_as_text(sheet.cell(row, source_description_column).value)
        article_code = (
            cell_as_text(sheet.cell(row, article_column).value) if article_column is not None else ""
        )
        output_title = sheet.cell(row, title_output_column).value
        output_description = sheet.cell(row, description_output_column).value

        has_output_title = not is_empty(output_title)
        has_output_description = not is_empty(output_description)
        # Лише повністю готовий рядок пропускається. Частково готовий рядок
        # доповнюється без перезапису наявного результату.
        if has_output_title and has_output_description:
            summary.skipped_completed += 1
            continue
        if not source_title or not source_description:
            summary.skipped_missing_source += 1
            continue

        try:
            missing_fields = []
            if not has_output_title:
                missing_fields.append("назва")
            if not has_output_description:
                missing_fields.append("опис")
            print(
                f"Рядок {row}: надсилаю запит до Gemini для поля {', '.join(missing_fields)}…",
                flush=True,
            )
            result = generator.generate(
                source_title,
                source_description,
                generate_title=not has_output_title,
                generate_description=not has_output_description,
                existing_title=cell_as_text(output_title),
                existing_description=cell_as_text(output_description),
                article_code=article_code,
            )
        except NoAvailableKeysError:
            raise
        except GenerationError as error:
            summary.failed_rows.append(row)
            log_gemini_error(
                "invalid_model_response",
                workbook_name=workbook_path.name,
                sheet_name=sheet_name,
                row=row,
                error_type=type(error).__name__,
                error_message=str(error),
            )
            print(f"Рядок {row}: пропущено через помилку — {error}")
            continue

        if not has_output_title:
            title_with_article = append_article_to_title(result["title"], article_code)
            if title_with_article != result["title"]:
                print("Gemini: назву скорочено з урахуванням артикула.", flush=True)
            validate_title(title_with_article)
            result["title"] = title_with_article
            sheet.cell(row, title_output_column).value = result["title"]
        if not has_output_description:
            sheet.cell(row, description_output_column).value = result["description"]
        # Помилка запису є критичною: не можна продовжувати з незбереженою книгою.
        save_workbook_atomically(workbook, workbook_path)
        summary.processed += 1
        print(f"Рядок {row}: збережено.")

    return summary


def ask_number(prompt: str, upper_bound: int) -> int:
    while True:
        raw = input(prompt).strip()
        try:
            value = int(raw)
        except ValueError:
            print("Введіть номер зі списку.")
            continue
        if 1 <= value <= upper_bound:
            return value
        print(f"Введіть число від 1 до {upper_bound}.")


def choose_workbook() -> Path:
    INPUT_DIR.mkdir(exist_ok=True)
    workbooks = sorted(
        path for path in INPUT_DIR.glob("*.xlsx") if not path.name.startswith("~$")
    )
    if not workbooks:
        raise ProcessorError(
            f"У теці {INPUT_DIR} немає .xlsx-файлів. Покладіть туди копію таблиці та запустіть програму знову."
        )
    print("Доступні Excel-файли:")
    for number, path in enumerate(workbooks, start=1):
        print(f"  {number}. {path.name}")
    return workbooks[ask_number("Оберіть номер файлу: ", len(workbooks)) - 1]


def choose_sheet(workbook: Workbook) -> Worksheet:
    print("Аркуші:")
    for number, name in enumerate(workbook.sheetnames, start=1):
        print(f"  {number}. {name}")
    return workbook[workbook.sheetnames[ask_number("Оберіть номер аркуша: ", len(workbook.sheetnames)) - 1]]


def choose_source_column(sheet: Worksheet, label: str, excluded: set[int] | None = None) -> int:
    excluded = excluded or set()
    columns = [
        column
        for column in range(1, sheet.max_column + 1)
        if column not in excluded and not is_empty(sheet.cell(HEADER_ROW, column).value)
    ]
    if not columns:
        raise ProcessorError("На першому рядку не знайдено колонок із заголовками.")
    print(f"Колонки для поля «{label}»:")
    for number, column in enumerate(columns, start=1):
        header = cell_as_text(sheet.cell(HEADER_ROW, column).value)
        print(f"  {number}. {get_column_letter(column)} — {header}")
    return columns[ask_number("Оберіть номер колонки: ", len(columns)) - 1]


def main() -> int:
    print("MONO Excel Processor")
    print(f"Модель: {MODEL_NAME}. Один запит Gemini створює і назву, і опис.\n")
    try:
        workbook_path = choose_workbook()
        workbook = load_workbook(workbook_path, read_only=True)
        try:
            sheet = choose_sheet(workbook)
            title_column = choose_source_column(sheet, "вихідна назва")
            description_column = choose_source_column(
                sheet, "вихідний опис", excluded={title_column}
            )
            article_column = choose_source_column(
                sheet,
                "артикул",
                excluded={title_column, description_column},
            )
            sheet_name = sheet.title
        finally:
            workbook.close()

        print(
            f"\nОбробка: {workbook_path.name}, аркуш «{sheet_name}», "
            f"колонки {get_column_letter(title_column)}, {get_column_letter(description_column)} "
            f"та {get_column_letter(article_column)}."
        )
        generator = GeminiGenerator(KeyPool())
        summary = run_processing(
            workbook_path,
            sheet_name,
            title_column,
            description_column,
            generator,
            article_column=article_column,
        )
    except (ProcessorError, OSError) as error:
        print(f"\nЗупинено: {error}")
        return 1

    print("\nГотово.")
    print(f"Успішно оброблено: {summary.processed}")
    print(f"Пропущено як уже готові: {summary.skipped_completed}")
    print(f"Пропущено без вихідних даних: {summary.skipped_missing_source}")
    if summary.failed_rows:
        rows = ", ".join(map(str, summary.failed_rows))
        print(f"Рядки з помилкою, які будуть повторені під час наступного запуску: {rows}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
