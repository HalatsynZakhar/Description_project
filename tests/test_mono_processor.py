from pathlib import Path
import json

from openpyxl import Workbook, load_workbook

import mono_processor
from mono_processor import (
    KeyPool,
    remove_decorative_tags,
    run_processing,
    shorten_title,
    source_description_for_model,
    validate_description,
    validate_title,
)


class FakeGenerator:
    def __init__(self) -> None:
        self.calls = 0
        self.requests: list[dict[str, object]] = []

    def generate(
        self,
        source_title: str,
        source_description: str,
        **options: object,
    ) -> dict[str, str]:
        self.calls += 1
        self.requests.append(options)
        result = {
            "title": "Навушники Acme X1 чорні (X1-BK)",
            "description": "<h5>Зручне прослуховування</h5><br><p>Бездротові навушники для щоденного використання.</p>",
        }
        return {
            field: value
            for field, value in result.items()
            if options.get(f"generate_{field}", True)
        }


def test_processor_saves_each_result_and_resumes(tmp_path: Path) -> None:
    path = tmp_path / "товари.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товари"
    sheet.append(["Стара назва", "Старий опис"])
    sheet.append(["Acme X1", "Бездротові навушники для щоденного використання."])
    sheet.append(["Без опису", None])
    workbook.save(path)

    generator = FakeGenerator()
    result = run_processing(path, "Товари", 1, 2, generator)
    assert result.processed == 1
    assert result.skipped_missing_source == 1
    assert generator.calls == 1

    saved = load_workbook(path)["Товари"]
    assert saved.cell(1, 3).value == "Назва MONO"
    assert saved.cell(1, 4).value == "Опис MONO"
    assert saved.cell(2, 3).value == "Навушники Acme X1 чорні (X1-BK)"

    rerun = run_processing(path, "Товари", 1, 2, generator)
    assert rerun.processed == 0
    assert rerun.skipped_completed == 1
    assert generator.calls == 1


def test_processor_fills_only_missing_mono_field(tmp_path: Path) -> None:
    path = tmp_path / "частково_готово.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товари"
    sheet.append(["Стара назва", "Старий опис", "Назва MONO", "Опис MONO"])
    sheet.append([
        "Acme X1",
        "Бездротові навушники для щоденного використання.",
        "Готова назва",
        None,
    ])
    workbook.save(path)

    generator = FakeGenerator()
    result = run_processing(path, "Товари", 1, 2, generator)
    assert result.processed == 1
    assert generator.requests == [{
        "generate_title": False,
        "generate_description": True,
        "existing_title": "Готова назва",
        "existing_description": "",
    }]

    saved = load_workbook(path)["Товари"]
    assert saved.cell(2, 3).value == "Готова назва"
    assert saved.cell(2, 4).value.startswith("<h5>")


def test_validation_rejects_links_and_bad_tags() -> None:
    validate_title("Кабель USB-C чорний")
    validate_description("<h5>Опис</h5><br><p>Надійний кабель.</p>", "Надійний кабель")

    try:
        validate_description("<img src=\"https://example.com/a.jpg\">", "Опис")
    except Exception as error:
        assert "URL" in str(error)
    else:
        raise AssertionError("Посилання має бути відхилено")


def test_decorative_tags_are_removed_before_mono_validation() -> None:
    description = remove_decorative_tags("<h5><strong>Опис</strong></h5><p>Текст</p>")
    assert description == "<h5>Опис</h5><p>Текст</p>"
    validate_description(description, "Текст")


def test_description_is_not_rejected_only_for_being_longer_than_source() -> None:
    description = "<p>" + ("Детальна характеристика. " * 100) + "</p>"
    validate_description(description, "Короткий опис")


def test_long_title_is_shortened_at_word_boundary() -> None:
    title = "Фігурка Marvel Веном із додатковою дуже довгою характеристикою кольору та особливостей моделі"
    shortened = shorten_title(title)
    assert len(shortened) <= 100
    assert shortened == shortened.rstrip()
    assert not shortened.endswith("характеристикою")


def test_source_sections_banned_by_mono_are_not_sent_to_model() -> None:
    source = "Особливості:\n* Світиться\nКомплектація:\n* Батарейки\nДодаткова інформація:\n* Режим"
    assert source_description_for_model(source) == "Особливості:\n* Світиться"


def test_tell_my_fortune_line_is_always_removed_before_model_request() -> None:
    source = "Особливості:\n* Режим Tell My Fortune для розваги\n* Світиться"
    assert source_description_for_model(source) == "Особливості:\n* Світиться"


def test_key_state_survives_restart(tmp_path: Path) -> None:
    keys_path = tmp_path / "keys.json"
    state_path = tmp_path / "keys_state.json"
    keys_path.write_text(
        '{"keys": [{"name": "one", "api_key": "secret", "enabled": true}]}', encoding="utf-8"
    )
    pool = KeyPool(keys_path, state_path)
    key = next(iter(pool.available()))
    pool.mark_failure(key, "rate_limited", "429 RESOURCE_EXHAUSTED", retry_after_seconds=60)

    restarted_pool = KeyPool(keys_path, state_path)
    assert list(restarted_pool.available()) == []


def test_gemini_error_log_has_diagnostics_without_secret(tmp_path: Path, monkeypatch) -> None:
    error_log = tmp_path / "logs" / "gemini_errors.jsonl"
    monkeypatch.setattr(mono_processor, "ERROR_LOG_PATH", error_log)
    mono_processor.log_gemini_error(
        "api_unavailable",
        key_name="project-a",
        category="transient",
        error_message="ReadTimeout",
    )

    record = json.loads(error_log.read_text(encoding="utf-8"))
    assert record["event"] == "api_unavailable"
    assert record["key_name"] == "project-a"
    assert "api_key" not in record
